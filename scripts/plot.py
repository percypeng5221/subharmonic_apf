#! /usr/bin/env python3
import rospkg
rospack = rospkg.RosPack()
packagePath = rospack.get_path('subharmAPF')

import yaml
with open(packagePath + '/config.yaml', 'r') as file: config = yaml.safe_load(file)
USE_SAMPLE = config['USE_SAMPLE']
USE_TRADITION = config['USE_TRADITION']

import sys
sys.path.insert(0, packagePath)

from openpyxl import load_workbook
wb1 = load_workbook(packagePath+'/data/result1.xlsx')

ws1 = wb1.active

x1 = []
y1 = []
laser1 = []

import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
# print(ws1.max_row)
for i in range (ws1.max_row-1):
    x1.append(ws1.cell(i+2, 1).value)
    y1.append(ws1.cell(i+2, 2).value)
    laser1.append(ws1.cell(i+2, 3).value)

fig = plt.figure(figsize=(15, 15)) 
plt.plot(x1, y1, '.', color='blue', label='robot1')
plt.gca().add_patch(Rectangle((0.75, 0.75), 0.5, 0.5, 
                    fill=True,
                    edgecolor = 'red',
                    facecolor = 'black',
                    linewidth=2))
plt.gca().add_patch(Rectangle((1, 2.6), 0.5, 0.5, 
                    fill=True,
                    edgecolor = 'red',
                    facecolor = 'black',
                    linewidth=2))
plt.gca().add_patch(Rectangle((2.25, 1.25), 0.5, 0.5, 
                    fill=True,
                    edgecolor = 'red',
                    facecolor = 'black',
                    linewidth=2))
plt.gca().add_patch(Rectangle((2.75, 2.75), 0.5, 0.5, 
                    fill=True,
                    edgecolor = 'red',
                    facecolor = 'black',
                    linewidth=2))
# plt.gca().add_patch(Rectangle((0.257537, 1.671751), 2, 0.1, 
#                     angle=-45,
#                     fill=True,
#                     edgecolor = 'red',
#                     facecolor = 'black',
#                     linewidth=2))
# plt.gca().add_patch(Rectangle((0.257537, 3.671751), 2, 0.1, 
#                     angle=-45,
#                     fill=True,
#                     edgecolor = 'red',
#                     facecolor = 'black',
#                     linewidth=2))
# plt.gca().add_patch(Rectangle((2.257537, 1.671751), 2, 0.1, 
#                     angle=-45,
#                     fill=True,
#                     edgecolor = 'red',
#                     facecolor = 'black',
#                     linewidth=2))
# plt.gca().add_patch(Rectangle((2.257537, 3.671751), 2, 0.1, 
#                     angle=-45,
#                     fill=True,
#                     edgecolor = 'red',
#                     facecolor = 'black',
#                     linewidth=2))
plt.legend(loc='best')  
plt.xlabel('x')
plt.ylabel('y')
# plt.ylim((-10, 0))
if USE_TRADITION and USE_SAMPLE: 
    plt.title("trajectory(traditional APF with sampling)")
    plt.savefig(packagePath+'/picture/trajectory(traditional APF with sampling).png')
if not USE_SAMPLE and USE_TRADITION:
    plt.title("trajectory(traditional APF without sampling)")
    plt.savefig(packagePath+'/picture/trajectory(traditional APF without sampling).png')
if USE_SAMPLE and not USE_TRADITION:
    plt.title("trajectory(harmonic APF with sampling)")
    plt.savefig(packagePath+'/picture/trajectory(harmonic APF with sampling).png')    
if not USE_SAMPLE and not USE_TRADITION:
    plt.title("trajectory(harmonic APF without sampling)")
    plt.savefig(packagePath+'/picture/trajectory(harmonic APF without sampling).png')    
plt.close(fig)

import numpy as np

plt.plot(range(len(laser1)), laser1, color='r', linestyle='solid', label='robot1')
plt.title('minimum distance to obstacle average = {average:.2f} min = {min:.2f}'.format(average = np.mean(laser1), min = min(laser1)))
plt.legend(loc='best')
plt.ylabel('distance/m')
plt.xlabel('timestep')

# 保存和展示
if USE_TRADITION and USE_SAMPLE: 
    plt.savefig(packagePath+'/picture/laser change(traditional APF with sampling).png')
elif not USE_SAMPLE and USE_TRADITION:
    plt.savefig(packagePath+'/picture/laser change(traditional APF without sampling).png')
elif USE_SAMPLE and not USE_TRADITION:
    plt.savefig(packagePath+'/picture/laser change(harmonic APF with sampling).png')    
elif not USE_SAMPLE and not USE_TRADITION:
    plt.savefig(packagePath+'/picture/laser change(harmonic APF without sampling).png')    
plt.close(fig)
    
    