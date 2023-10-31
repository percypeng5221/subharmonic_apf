#! /usr/bin/env python3
import sys
import rospkg
import yaml
rospack = rospkg.RosPack()
packagePath = rospack.get_path('subharmonic_apf')
sys.path.insert(0, packagePath+'/scripts')

from openpyxl import Workbook,load_workbook
wb = Workbook()
ws = wb.active

import rospy
from Control import *
from Lidar import *
import field
import control
import modify_control
from sensor_msgs.msg import LaserScan 
from geometry_msgs.msg import Twist
from nav_msgs.msg import Odometry
from gazebo_msgs.msg import ModelState
from tf.transformations import quaternion_from_euler, euler_from_quaternion
import math

with open(packagePath + '/config.yaml', 'r') as file: config = yaml.safe_load(file)
# Initial and goal positions
X_INIT = config['INIT_POSITIONS_X']['first']
Y_INIT = config['INIT_POSITIONS_Y']['first']
THETA_INIT = config['INIT_POSITIONS_THETA']['first']

X_GOAL = config['GOAL_POSITIONS_X']['first']
Y_GOAL = config['GOAL_POSITIONS_Y']['first']
THETA_GOAL = config['GOAL_POSITIONS_THETA']['first']
V_K = config['V_K']
DISTANCE = config['SAMPLE_DISTANCE']
NUM = config['NUM_POINTS']
USE_SAMPLE = config['USE_SAMPLE']
ONLY_PLAN = config['ONLY_PLAN']
REACH_GOAL = config['REACH_GOAL']

odomMsg1 = Odometry()
msgScan1 = LaserScan()

def odomCallback1(odometry):
    global odomMsg1
    odomMsg1 = odometry
    
def scanCallback1(laserScan):
    global msgScan1
    msgScan1 = laserScan
    
def resetPos(setPosPub, x, y):
    checkpoint = ModelState()
    checkpoint.model_name = "tb3_1"
    checkpoint.pose.position.x = x
    checkpoint.pose.position.y = y
    checkpoint.pose.position.z = 0.0
    [x_q,y_q,z_q,w_q] = quaternion_from_euler(0.0, 0.0, math.radians(THETA_INIT))
    checkpoint.pose.orientation.x = x_q
    checkpoint.pose.orientation.y = y_q
    checkpoint.pose.orientation.z = z_q
    checkpoint.pose.orientation.w = w_q
    checkpoint.twist.linear.x = 0.0
    checkpoint.twist.linear.y = 0.0
    checkpoint.twist.linear.z = 0.0
    checkpoint.twist.angular.x = 0.0
    checkpoint.twist.angular.y = 0.0
    checkpoint.twist.angular.z = 0.0
    setPosPub.publish(checkpoint)

if __name__ == '__main__':
    
    rospy.init_node('control_node', anonymous = True)
    velPub = rospy.Publisher('/tb3_1/cmd_vel', Twist, queue_size = 10)
    odomSub1 = rospy.Subscriber('/tb3_1/odom', Odometry, odomCallback1, queue_size=10)
    scanSub = rospy.Subscriber('/tb3_1/scan', LaserScan, scanCallback1, queue_size=10)
    setPosPub = rospy.Publisher('/gazebo/set_model_state', ModelState, queue_size = 10)
    rospy.sleep(0.2)
    while (field.getDistance(odomMsg1.pose.pose.position.x, X_INIT, odomMsg1.pose.pose.position.y, Y_INIT) > REACH_GOAL): 
        resetPos(setPosPub, X_INIT, Y_INIT)

    rospy.sleep(0.2)
    row = 1
    while field.getDistance(odomMsg1.pose.pose.position.x, X_GOAL, odomMsg1.pose.pose.position.y, Y_GOAL) > REACH_GOAL:
        # rospy.sleep(0.2)
        time1 = rospy.get_time()
        forceObs, angleObs = field.lasarField(msgScan1, odomMsg1, False)
        forceTarget, angleTarget = field.targetField(X_GOAL, Y_GOAL, odomMsg1)
        forceX = forceObs * math.cos(angleObs) + forceTarget * math.cos(angleTarget) 
        forceY = forceObs * math.sin(angleObs) + forceTarget * math.sin(angleTarget)
                
        forceMagnitude = (forceX ** 2 + forceY ** 2) ** 0.5
        forceAngle = math.atan2(forceY, forceX)
        tempx = odomMsg1.pose.pose.position.x + forceX*V_K
        tempy = odomMsg1.pose.pose.position.y + forceY*V_K
        temp_quaternion = [odomMsg1.pose.pose.orientation.x, odomMsg1.pose.pose.orientation.y,
                           odomMsg1.pose.pose.orientation.z, odomMsg1.pose.pose.orientation.w]
        temp_orientation = euler_from_quaternion(temp_quaternion)
        
        if USE_SAMPLE: 
            targetDirection = math.atan2(Y_GOAL - odomMsg1.pose.pose.position.y, X_GOAL - odomMsg1.pose.pose.position.x)
            # print(targetDirection, temp_orientation[2])
            xPoints, yPoints = field.generatePoints(tempx, tempy, targetDirection, DISTANCE, NUM) # temp_orientation[2]
            values = field.evaluateLaser(msgScan1, odomMsg1, X_GOAL, Y_GOAL, xPoints, yPoints, plotFlag=False)
            tempx = xPoints[values.index(min(values))]
            tempy = yPoints[values.index(min(values))]
            
        row += 1
        time2 = rospy.get_time()
        if ONLY_PLAN: 
            resetPos(setPosPub, tempx, tempy)
            distance, _ = lidarScan(msgScan1)
            ws.cell(row, 1).value = tempx
            ws.cell(row, 2).value = tempy
            ws.cell(row, 3).value = min(distance)
            rospy.sleep(0.2 + time1 - time2)
        else: 
            while(time2 - time1 < 0.2):
                velMsg = modify_control.move(odomMsg1, tempx, tempy)
                velPub.publish(velMsg)
                distance, _ = lidarScan(msgScan1)
                ws.cell(row, 1).value = tempx
                ws.cell(row, 2).value = tempy
                ws.cell(row, 3).value = min(distance)
                time2 = rospy.get_time()
                # print(velMsg.angular.z, velMsg.linear.x)
                # print(time2 - time1)
    ws.cell(1, 1).value = "x"
    ws.cell(1, 2).value = "y"
    ws.cell(1, 3).value = "distance to obstacle" 
    wb.save(packagePath+'/data/result1 traditional, without sampling.xlsx')
    
    # while (field.getDistance(odomMsg1.pose.pose.position.x, X_INIT, odomMsg1.pose.pose.position.y, Y_INIT) > 0.1): 
    #     resetPos(setPosPub, X_INIT, Y_INIT)

        
                 
    
    
        
    
    
